import openpyxl
import json
from datetime import datetime

def cell_to_str(val):
    if val is None:
        return ""
    if isinstance(val, datetime):
        return val.strftime("%Y-%m-%d %H:%M:%S")
    if isinstance(val, float):
        if val == int(val):
            return str(int(val))
        return str(val)
    return str(val)

wb = openpyxl.load_workbook('D:/NOVEL/NOVE.xlsx', data_only=True)
seed_data = {}

library_defs = {
    '热梗素材库': {
        'prefix': 'RG', 'sheet': '热梗素材库',
        'headers': ['编号','类型标签','来源平台','核心冲突点','情感层次','改编方向','甜虐指数','甜虐平衡指数','适配角色关系','适用章节','道具符号','禁忌红线','多平台热度','伏笔需求','读者预期','风险提示','适用情节']
    },
    '冲突素材库': {
        'prefix': 'CT', 'sheet': '冲突素材库',
        'headers': ['编号','冲突类型','来源平台','现实原型','核心矛盾','情绪曲线','适配角色关系','适用场景','风险等级','改编方向','改编案例','爆发场景','解决方式','情绪价值','伏笔需求','适用情节']
    },
    '钩子素材库': {
        'prefix': 'GZ', 'sheet': '钩子素材库',
        'headers': ['编号','钩子类型','埋设位置','类型标签','引爆章节','核心元素','埋设手法','情感导向','回收周期','关联反转','钩子密度','伏笔要求','修改方向','风险提示','适用情节']
    },
    '反转素材库': {
        'prefix': 'FZ', 'sheet': '反转素材库',
        'headers': ['编号','反转类型','铺垫线索','爆发章节','情感冲击','逻辑校验','表面呈现','真相揭露','伏笔设计','改编案例','现实原型','适用题材','风险屏障','适用情节']
    },
    '人设基因库': {
        'prefix': 'RS', 'sheet': '人设基因库',
        'headers': ['编号','角色定位','性格标签','标签','职业/身份','情感障碍','专属物品','标志动作','感情线伏笔','隐秘关联','荷尔蒙触发点','致命性缺点','救赎开关','禁忌边界','适用情节']
    },
    '场景库': {
        'prefix': 'CJ', 'sheet': '场景兵器库',
        'headers': ['编号','场景类型','视觉焦点','听觉细节','触觉意象','嗅觉记忆','味觉隐喻','情感曲线','冲突触发点','伏笔回收点','适用情节','数据参考','禁忌提示','情感强度']
    },
    '法律风险库': {
        'prefix': 'FL', 'sheet': '法律风险库',
        'headers': ['编号','风险类型','法律依据','触发条件','后果严重性','现实原型','改编方案','预防措施','戏剧化技巧','适用情节']
    },
    '词汇库': {
        'prefix': 'CH', 'sheet': '词汇库提取',
        'headers': ['编号','分类','子类','核心词汇','强度','通感转化示例','适用题材','关联情绪','出处']
    },
    '情绪库': {
        'prefix': 'QX', 'sheet': '情绪库模板',
        'headers': ['编号','核心情绪','强度等级','生理反应','微表情编码','行为映射','对话特征','适用场景','禁忌误用','案例来源']
    },
    '景色库': {
        'prefix': 'JS', 'sheet': '景色库模板',
        'headers': ['编号','时空坐标','光学描写','声学描写','嗅觉层次','触觉反馈','动态元素','数据层','时代标签','关联色卡']
    },
    '动作库': {
        'prefix': 'DZ', 'sheet': '动作库模板',
        'headers': ['编号','场景类型','动作分级','主体动作','连带反应','隐喻意义','节奏值','禁忌组合','经典案例']
    },
    '对话库': {
        'prefix': 'DH', 'sheet': '对话库模板',
        'headers': ['编号','冲突类型','表层对话','潜台词','动作锚点','声调标记','信息密度','权力关系','出处章节']
    },
    '金句库': {
        'prefix': 'JJ', 'sheet': '金句库',
        'headers': ['编号','金句内容','类型标签','声韵结构','隐喻密度','跨库冲击力','关联热梗','适用情节']
    },
    '幽默素材库': {
        'prefix': 'HM', 'sheet': '幽默',
        'headers': ['编号','评论原文','核心笑点','幽默类型','可复用结构','优化建议','情绪强度','适用场景','高频关键词','禁忌提示','参考作品','创作周期','对应章节','埋梗位置','黑色幽默指数','冷笑话指数']
    },
}

for lib_id, lib_def in library_defs.items():
    sheet_name = lib_def['sheet']
    if sheet_name not in wb.sheetnames:
        print(f"WARNING: Sheet '{sheet_name}' not found, skipping {lib_id}")
        continue
    ws = wb[sheet_name]
    items = []
    headers = lib_def['headers']
    for row_idx in range(3, ws.max_row + 1):
        row_data = {}
        has_data = False
        for col_idx, field in enumerate(headers, 1):
            val = cell_to_str(ws.cell(row=row_idx, column=col_idx).value)
            row_data[field] = val
            if val and field == '编号':
                has_data = True
        if has_data and row_data.get('编号'):
            id_val = row_data['编号']
            if id_val.startswith(('1.','2.','3.','4.','5.','6.','7.','8.','9.','10.','11.','12.','13.','14.','15.')):
                continue
            items.append(row_data)
    seed_data[lib_id] = {
        'name': lib_id,
        'prefix': lib_def['prefix'],
        'headers': headers,
        'items': items
    }
    print(f"{lib_id}: {len(items)} items")

# Character card
ws_char = wb['角色卡']
char_fields = []
for row in ws_char.iter_rows(min_row=2, max_row=ws_char.max_row, values_only=False):
    seq = cell_to_str(row[0].value)
    category = cell_to_str(row[1].value)
    field_name = cell_to_str(row[2].value)
    hint = cell_to_str(row[3].value)
    example = cell_to_str(row[4].value)
    if field_name:
        last_cat = char_fields[-1]['category'] if char_fields else ''
        char_fields.append({
            'seq': seq,
            'category': category if category else last_cat,
            'field': field_name,
            'hint': hint,
            'example': example
        })

example_char = {}
for row in ws_char.iter_rows(min_row=2, max_row=ws_char.max_row, values_only=False):
    field_name = cell_to_str(row[2].value)
    example = cell_to_str(row[4].value)
    if field_name and example:
        example_char[field_name] = example
if example_char:
    example_char['编号'] = 'CHAR-001'
    example_char['姓名'] = example_char.get('姓名', '祝余嘉')
    seed_data['角色卡'] = {
        'name': '角色卡', 'prefix': 'CHAR',
        'fields': char_fields, 'items': [example_char]
    }
else:
    seed_data['角色卡'] = {
        'name': '角色卡', 'prefix': 'CHAR',
        'fields': char_fields, 'items': []
    }
print(f"\n角色卡: {len(char_fields)} fields, {len(seed_data['角色卡']['items'])} example characters")

# 扫榜
ws_scan = wb['扫榜']
scan_items = []
for row in ws_scan.iter_rows(min_row=2, max_row=min(ws_scan.max_row, 100), values_only=False):
    row_data = {
        '章节': cell_to_str(row[0].value),
        '章节标题': cell_to_str(row[1].value),
        '内容提要': cell_to_str(row[2].value),
        '字数': cell_to_str(row[3].value),
        '点击': cell_to_str(row[4].value),
        '更新时间': cell_to_str(row[5].value),
        '文章名': cell_to_str(row[6].value),
        '作者': cell_to_str(row[7].value),
        '标签': cell_to_str(row[8].value),
    }
    if row_data['章节标题'] or row_data['内容提要']:
        scan_items.append(row_data)
seed_data['扫榜'] = {
    'name': '扫榜',
    'headers': ['章节','章节标题','内容提要','字数','点击','更新时间','文章名','作者','标签'],
    'items': scan_items
}
print(f"扫榜: {len(scan_items)} items (sampled)")

# 大纲
ws_outline = wb['大纲']
outline_headers = ['故事阶段','大事件','章节名称','事业线','感情线','时间','事件内容','主要人物','看点','问题','流程','伏笔','知识','写后梗概']
outline_items = []
for row in ws_outline.iter_rows(min_row=6, max_row=ws_outline.max_row, values_only=False):
    row_data = {}
    for i, h in enumerate(outline_headers):
        if i < len(row):
            row_data[h] = cell_to_str(row[i].value)
    if row_data.get('故事阶段'):
        outline_items.append(row_data)
seed_data['大纲'] = {
    'name': '大纲', 'headers': outline_headers, 'items': outline_items
}
print(f"大纲: {len(outline_items)} items")

output_path = 'E:/写作/清钰的写作库/1.小说创作/小说创作平台/js/data/seed-data.json'
with open(output_path, 'w', encoding='utf-8') as f:
    json.dump(seed_data, f, ensure_ascii=False, indent=2)
print(f"\nSeed data saved to: {output_path}")
print(f"Total libraries: {len(seed_data)}")
